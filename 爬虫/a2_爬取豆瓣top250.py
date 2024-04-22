import requests
import xlwt
from bs4 import BeautifulSoup
import re
import openpyxl
from openpyxl.styles import Font

datas = []
page = 0

type_words = [
    "动作", "喜剧", "剧情", "悬疑", "惊悚", "科幻", "奇幻", "恐怖",
    "爱情", "战争", "历史", "音乐", "西部", "青春", "家庭", "儿童",
    "体育", "传记", "冒险", "犯罪", "纪录片", "歌舞", "情色", "灾难",
    "黑色电影", "治愈系", "武侠", "军事", "校园", "励志", "侦探",
    "超级英雄", "动画", "网络", "实验", "独立", "短片", "长片", "默片",
    "同性"
]


def get_str(s: int, e: int, ch: str):
    result = ''
    for n in range(s + 1, e):
        if n == e - 1:
            result = result + str(info[n])
            break
        result = result + str(info[n]) + ch
    return result


times = 1
while True:  # 爬不到完整的数据一直爬
    print(f"*********第{times}次")
    while page <= 225:
        # 1、爬取数据
        url = f'https://movie.douban.com/top250?start={page}&filter='
        print(url)
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0'}
        response = requests.get(url, headers=headers)

        soup = BeautifulSoup(response.text, "html.parser")
        movies = soup.find_all("div", class_="item")
        for movie in movies:
            data = []
            # 播放地址
            play_link = str(movie.find("a")["href"])
            data.append(play_link)
            # 海报地址
            poster = str(movie.find("img")["src"])
            data.append(poster)
            # 电影名称
            title = movie.find("span", class_="title").text
            data.append(title)

            info = movie.find("div", class_="bd").find("p").text.strip()
            info = re.compile(r'''[\u4e00-\u9fa5\d]+''').findall(info)
            start = info.index('导演')
            try:
                end = info.index('主演')
            except ValueError:
                try:
                    end = info.index('主')
                except ValueError:
                    end = info.index(str(re.compile(r'''\d+''').findall(info.__str__())[0]).strip("\n"))

            director = get_str(start, end, "·")
            data.append(director)

            start = end
            date = str(re.compile(r'''\d+''').findall(info.__str__())[0]).strip("\n")
            end = info.index(date)
            main_actors = get_str(start, end, "、")
            # 电影导演
            data.append(main_actors)
            # 上映时间
            data.append(date)
            start = end
            end += 2
            while True:
                try:
                    type_words.index(info[end])
                    break
                except ValueError:
                    end += 1
            # 发行国家
            publish_country = get_str(start, end, "、")
            data.append(publish_country)
            # 电影类型
            start = end - 1
            end = len(info)
            movie_type = get_str(start, end, "、")
            data.append(movie_type)
            # 电影评分
            start = end
            rating_rating_count = str(movie.find("div", class_="star").text).strip().split("\n\n")
            rating = rating_rating_count[0]
            data.append(rating)
            # 评分人数
            rating_count = rating_rating_count[1]
            data.append(rating_count)

            # 电影主题
            try:
                theme_point = movie.find("span", class_="inq").text
            except AttributeError:
                theme_point = ""

            data.append(theme_point)

            datas.append(data)
        page += 25
        # time.sleep(3)
    print(f"一共爬取到{len(datas)}条数据")
    if len(datas) >= 200:
        break
    else:
        datas = []
        page = 0
        times += 1

book = xlwt.Workbook(encoding='utf-8', style_compression=0)
# 创建一个Excel Workbook表格，用于存储Excel数据
sheet = book.add_sheet("豆瓣电影TOP250", cell_overwrite_ok=True)
# 在Workbook中添加一个名为“豆瓣电影TOP250”的工作表

col = (
    '播放地址', '电影海报', '电影名称', '导演', '主演', '上映时间', '出品国家', '电影类型', '评分', '评价人数(人)',
    '灵魂动人点',
)

# 创建工作簿
workbook = openpyxl.Workbook()

# 选择工作表
worksheet = workbook.active

# 设置列宽
for c in range(ord('A'), ord('K') + 1):
    worksheet.column_dimensions[chr(c)].width = 20

# 设置行高
worksheet.row_dimensions[0].height = 40
for row in range(1, 252):
    worksheet.row_dimensions[row].height = 30

# 写入数据
# 写入表头数据
index = 1
for element in col:
    cell = worksheet.cell(row=1, column=index)
    cell.value = f"{element}"
    # 格式化
    cell.font = Font(size=22)
    index += 1

# 从第二行开始逐行写入爬取到的电影数据
row_num = 2
col_num = 1
for val in datas:
    for element in val:
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = f"{element}"
        col_num += 1
    row_num += 1
    col_num = 1

# 保存文件
workbook.save('top250_.xlsx')
